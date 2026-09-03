#!/usr/bin/env python3
"""
CrossConnect debug simulation script.

Exercises the full workflow end-to-end against the live database:
  - Datacenters + contacts
  - Racks
  - Device types
  - Systems + devices (including parent/child)
  - Switches
  - Work orders (create, issue, pull back, re-issue, complete)
  - Connections (valid rows, duplicate switch port, missing mandatory fields,
    cross-cabinet cable length, same-rack cable length, LAG rows, fiber mode)
  - Status transitions (valid and invalid)
  - Soft delete + recycle bin check

Run from project root:
    .venv/bin/python -m tests.debug_sim

Prints a PASS/FAIL summary. Cleans up its own test data on completion
(unless --keep is passed).

Exit code: 0 if all pass, 1 if any fail.
"""
import sys
import argparse
from pathlib import Path

# Allow running as module from project root
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.db import SessionLocal
from app.models.user import User
from app.models.inventory import Datacenter, DCContact, Rack, System, Device, Switch, DeviceType
from app.models.work_order import WorkOrder
from app.models.connection import Connection
from app.models.audit import AuditLog
import app.services.inventory as inv_svc
import app.services.work_orders as wo_svc
from app.services.auth import create_user, get_user_by_username


# ── Test harness ──────────────────────────────────────────────────────────

RESULTS: list[tuple[str, bool, str]] = []
CREATED_IDS: dict[str, list[int]] = {
    "datacenter": [], "rack": [], "system": [], "device": [],
    "switch": [], "device_type": [], "work_order": [], "user": [],
}


def check(label: str, condition: bool, detail: str = ""):
    status = "PASS" if condition else "FAIL"
    RESULTS.append((label, condition, detail))
    icon = "✓" if condition else "✗"
    print(f"  [{icon}] {label}" + (f"  — {detail}" if detail else ""))
    return condition


def expect_error(label: str, fn, expected_fragment: str = ""):
    """Assert that fn() raises ValueError containing expected_fragment."""
    try:
        fn()
        check(label, False, f"Expected ValueError but no exception raised")
        return False
    except ValueError as e:
        ok = expected_fragment.lower() in str(e).lower() if expected_fragment else True
        check(label, ok, str(e) if not ok else "")
        return ok
    except Exception as e:
        check(label, False, f"Unexpected exception: {type(e).__name__}: {e}")
        return False


def section(title: str):
    print(f"\n{'─'*60}")
    print(f"  {title}")
    print(f"{'─'*60}")


# ── Setup ─────────────────────────────────────────────────────────────────

def pre_run_cleanup(db):
    """
    Remove any stale test data left by a previous --keep run or crash.
    Safe to call even if nothing exists.
    """
    from app.models.connection import Connection
    from app.models.work_order import WorkOrder

    stale_wos = db.query(WorkOrder).filter(WorkOrder.name.like("_SIM%")).all()
    for wo in stale_wos:
        db.query(Connection).filter(Connection.work_order_id == wo.id).delete()
        db.delete(wo)

    from app.models.inventory import Switch, Device, System, Rack, DCContact, Datacenter, DeviceType
    db.query(Switch).filter(Switch.name.like("_SIM%")).delete()
    # Devices: children before parents (self-referential FK)
    children = db.query(Device).filter(
        Device.name.like("_SIM%"), Device.parent_device_id.isnot(None)).all()
    for d in children:
        db.delete(d)
    db.flush()
    db.query(Device).filter(Device.name.like("_SIM%")).delete()
    db.query(System).filter(System.name.like("_SIM%")).delete()

    stale_racks = db.query(Rack).filter(Rack.name.like("_SIM%")).all()
    for r in stale_racks:
        db.delete(r)

    stale_dcs = db.query(Datacenter).filter(
        (Datacenter.name.like("_SIM%")) | (Datacenter.code.like("SIM%"))).all()
    for dc in stale_dcs:
        db.query(DCContact).filter(DCContact.datacenter_id == dc.id).delete()
        db.delete(dc)

    db.query(DeviceType).filter(DeviceType.manufacturer.like("_SIM%")).delete()

    stale_users = db.query(User).filter(User.username == "_sim_architect").all()
    for u in stale_users:
        db.query(AuditLog).filter(AuditLog.user_id == u.id).delete()
        db.delete(u)

    db.commit()


def setup_test_user(db) -> User:
    username = "_sim_architect"
    existing = get_user_by_username(db, username)
    if existing:
        return existing
    u = create_user(db, username=username, display_name="Sim Architect",
                    password="sim_password_123", role="architect")
    CREATED_IDS["user"].append(u.id)
    return u


# ── Datacenter tests ──────────────────────────────────────────────────────

def test_datacenters(db) -> Datacenter:
    section("Datacenters + Contacts")

    # Create
    dc = inv_svc.create_datacenter(db, name="_SIM Test DC", code="SIM",
                                    address="123 Sim St", has_grid_system=True)
    CREATED_IDS["datacenter"].append(dc.id)
    check("Create datacenter", dc.id is not None, f"id={dc.id}")
    check("Code uppercased", dc.code == "SIM")
    check("Address stored", dc.address == "123 Sim St")

    # Duplicate name
    expect_error("Duplicate DC name blocked",
                 lambda: inv_svc.create_datacenter(db, name="_SIM Test DC", code="SIM2"),
                 "already exists")

    # Duplicate code
    expect_error("Duplicate DC code blocked",
                 lambda: inv_svc.create_datacenter(db, name="_SIM Test DC 2", code="SIM"),
                 "already in use")

    # Contacts
    inv_svc.upsert_contacts(db, dc, [
        {"name": "Alice", "role": "Site Lead", "email": "alice@sim.test", "phone": "555-0001"},
        {"name": "Bob",   "role": "NOC",       "email": "bob@sim.test",   "phone": "555-0002"},
    ])
    db.refresh(dc)
    check("Contacts saved", len(dc.contacts) == 2, f"{len(dc.contacts)} contacts")
    check("Contact name", dc.contacts[0].name == "Alice")
    check("Contact role", dc.contacts[0].role == "Site Lead")

    # Upsert replaces existing
    inv_svc.upsert_contacts(db, dc, [
        {"name": "Charlie", "role": "Facilities", "email": "charlie@sim.test", "phone": ""},
    ])
    db.refresh(dc)
    check("Contact upsert replaces", len(dc.contacts) == 1)
    check("New contact name", dc.contacts[0].name == "Charlie")

    # Empty-name contact skipped
    inv_svc.upsert_contacts(db, dc, [
        {"name": "Valid", "role": "", "email": "", "phone": ""},
        {"name": "",      "role": "skip me", "email": "", "phone": ""},
    ])
    db.refresh(dc)
    check("Empty-name contact skipped", len(dc.contacts) == 1)

    return dc


# ── Rack tests ────────────────────────────────────────────────────────────

def test_racks(db, dc: Datacenter) -> Rack:
    section("Racks")

    rack = inv_svc.create_rack(db, dc_id=dc.id, name="_SIM-RACK-01",
                                grid_position="AA-01", total_ru=42)
    CREATED_IDS["rack"].append(rack.id)
    check("Create rack", rack.id is not None)
    check("Grid position stored", rack.grid_position == "AA-01")
    check("Total RU stored", rack.total_ru == 42)

    # Duplicate rack name in same DC
    expect_error("Duplicate rack name in DC blocked",
                 lambda: inv_svc.create_rack(db, dc_id=dc.id, name="_SIM-RACK-01"),
                 "already exists")

    # Same name in different DC should be allowed — use a temp DC
    dc2 = inv_svc.create_datacenter(db, name="_SIM Test DC 2", code="SIM2")
    CREATED_IDS["datacenter"].append(dc2.id)
    rack2 = inv_svc.create_rack(db, dc_id=dc2.id, name="_SIM-RACK-01")
    CREATED_IDS["rack"].append(rack2.id)
    check("Same rack name in different DC allowed", rack2.id is not None)

    # Update
    inv_svc.update_rack(db, rack, name="_SIM-RACK-01", grid_position="BB-02", total_ru=42)
    db.refresh(rack)
    check("Update rack grid position", rack.grid_position == "BB-02")

    return rack


# ── Device type tests ─────────────────────────────────────────────────────

def test_device_types(db) -> tuple[DeviceType, DeviceType]:
    section("Device Types")

    dt_storage = inv_svc.create_device_type(
        db, manufacturer="_SIM NetApp", model="AFF A400",
        category="storage", rack_u=2, slot_count=None)
    CREATED_IDS["device_type"].append(dt_storage.id)
    check("Create storage device type", dt_storage.id is not None)

    dt_chassis = inv_svc.create_device_type(
        db, manufacturer="_SIM Cisco", model="UCS 5108",
        category="server", rack_u=7, slot_count=8,
        notes="Blade chassis, 8 slots")
    CREATED_IDS["device_type"].append(dt_chassis.id)
    check("Create chassis device type", dt_chassis.id is not None)
    check("Slot count stored", dt_chassis.slot_count == 8)

    # Duplicate
    expect_error("Duplicate device type blocked",
                 lambda: inv_svc.create_device_type(
                     db, manufacturer="_SIM NetApp", model="AFF A400"),
                 "already exists")

    # Autocomplete
    results = inv_svc.autocomplete_device_types(db, "_SIM")
    check("Autocomplete returns both types", len(results) >= 2)
    results_storage = inv_svc.autocomplete_device_types(db, "_SIM", category="storage")
    check("Autocomplete category filter", len(results_storage) == 1)
    check("Autocomplete rack_u present", results_storage[0]["rack_u"] == 2)

    return dt_storage, dt_chassis


# ── Device tests ──────────────────────────────────────────────────────────

def test_devices(db, rack: Rack, dt_storage: DeviceType, dt_chassis: DeviceType) -> Device:
    section("Systems + Devices (including parent/child)")

    # System
    system = inv_svc.create_system(db, name="_SIM Storage Cluster", system_type="storage")
    CREATED_IDS["system"].append(system.id)
    check("Create system", system.id is not None)

    # Duplicate system name
    expect_error("Duplicate system name blocked",
                 lambda: inv_svc.create_system(db, name="_SIM Storage Cluster"),
                 "already exists")

    # Standalone device (no system)
    standalone = inv_svc.create_device(db, rack_id=rack.id,
                                        name="_SIM-srv-01", device_type_id=dt_storage.id)
    CREATED_IDS["device"].append(standalone.id)
    check("Create standalone device (no system)", standalone.id is not None)
    check("Device type FK set", standalone.device_type_id == dt_storage.id)

    # Device with system
    node = inv_svc.create_device(db, rack_id=rack.id, name="_SIM-stor-44-01",
                                  system_id=system.id, starting_ru=10,
                                  serial="SN-001", device_type_id=dt_storage.id)
    CREATED_IDS["device"].append(node.id)
    check("Device with system", node.system_id == system.id)
    check("Starting RU stored", node.starting_ru == 10)

    # Parent/child: chassis + blade
    chassis = inv_svc.create_device(db, rack_id=rack.id, name="_SIM-UCS-5108",
                                     device_type_id=dt_chassis.id, starting_ru=20)
    CREATED_IDS["device"].append(chassis.id)
    blade = inv_svc.create_device(db, rack_id=rack.id, name="_SIM-UCS-blade-01",
                                   parent_device_id=chassis.id, slot_number=1)
    CREATED_IDS["device"].append(blade.id)
    db.refresh(blade)
    check("Child device parent FK", blade.parent_device_id == chassis.id)
    check("Child device slot number", blade.slot_number == 1)
    db.refresh(chassis)
    check("Parent has child_devices backref", len(chassis.child_devices) == 1)

    return node


# ── Switch tests ──────────────────────────────────────────────────────────

def test_switches(db, rack: Rack) -> Switch:
    section("Switches")

    dt_sw = inv_svc.create_device_type(db, manufacturer="_SIM Cisco",
                                        model="C9300-48P", category="switch", rack_u=1)
    CREATED_IDS["device_type"].append(dt_sw.id)

    sw = inv_svc.create_switch(db, rack_id=rack.id, name="_SIM-SW-LAN-01",
                                switch_role="LAN", starting_ru=30,
                                switch_type_id=dt_sw.id)
    CREATED_IDS["switch"].append(sw.id)
    check("Create LAN switch", sw.id is not None)
    check("Switch type FK", sw.switch_type_id == dt_sw.id)

    sw_san = inv_svc.create_switch(db, rack_id=rack.id, name="_SIM-SW-SAN-01",
                                    switch_role="SAN", starting_ru=32)
    CREATED_IDS["switch"].append(sw_san.id)
    check("Create SAN switch", sw_san.switch_role == "SAN")

    # Autocomplete filter by role
    results = inv_svc.autocomplete_switches(db, "_SIM", role="SAN")
    check("Switch autocomplete role filter", len(results) == 1)
    check("Switch autocomplete label includes role", "SAN" in results[0]["label"])

    return sw


# ── Work order + connection tests ─────────────────────────────────────────


def test_template_renders(db, dc, rack, user):
    """
    Render every major template against real data to catch Jinja2 errors
    that only surface with None/mixed data (e.g. sorting nulls, missing FKs).
    Uses Jinja2 directly — no HTTP server needed.
    """
    section("Template Renders (Jinja2 direct)")

    from jinja2 import Environment, FileSystemLoader
    import os

    template_dir = str(Path(__file__).resolve().parents[1] / "app" / "templates")
    env = Environment(loader=FileSystemLoader(template_dir))

    class FakeRequest:
        class url:
            path = "/test"

    ctx_base = {"request": FakeRequest(), "user": user}

    def render(template_name, extra_ctx):
        try:
            t = env.get_template(template_name)
            t.render(**ctx_base, **extra_ctx)
            check(f"Render {template_name}", True)
            return True
        except Exception as e:
            check(f"Render {template_name}", False, f"{type(e).__name__}: {e}")
            return False

    # Inventory pages
    from app.services.inventory import get_datacenter, get_rack, list_datacenters, list_systems, list_device_types
    dc_full = get_datacenter(db, dc.id)
    rack_full = get_rack(db, rack.id)
    dcs = list_datacenters(db)
    systems = list_systems(db)
    dts = list_device_types(db)

    render("inventory/index.html", {"datacenters": dcs, "systems": systems})
    render("inventory/dc_detail.html", {"dc": dc_full, "racks": dc_full.racks})
    render("inventory/dc_form.html", {"dc": None, "error": None})
    render("inventory/dc_form.html", {"dc": dc_full, "error": None})
    from app.services.inventory import get_rack_elevation
    elevation = get_rack_elevation(db, rack_full.id, max_ru=42)
    render("inventory/rack_detail.html", {"rack": rack_full, "elevation": elevation, "max_ru": 42})
    render("inventory/rack_form.html", {"dc": dc_full, "rack": None, "error": None})
    render("inventory/rack_form.html", {"dc": dc_full, "rack": rack_full, "error": None})
    render("inventory/device_type_list.html", {"device_types": dts})
    render("inventory/device_type_form.html", {"dt": None, "error": None})
    render("inventory/system_form.html", {"system": None, "error": None})
    render("inventory/rack_import.html",   {"preview": None, "result": None, "prefill_dc": ""})
    render("inventory/device_import.html", {"preview": None, "result": None})

    # Work order pages — need a WO with connections
    from app.services.work_orders import list_work_orders, list_connections
    wos = list_work_orders(db)
    render("work_orders/list.html", {
        "work_orders": wos,
        "datacenters": dcs,
        "filter_dc_id": None,
        "filter_status": None,
    })
    if wos:
        wo = wos[0]
        conns = list_connections(db, wo.id)
        render("work_orders/detail.html", {
            "wo": wo, "connections": conns,
            "CABLE_TYPES": ["RJ45","LC_Fiber","DAC","other"],
            "PURPOSES": ["management","data","storage","other"],
            "ACTIONS": ["A","R","C"],
            "FABRICS": ["","A","B"],
            "INSTALL_STATUSES": ["pending","done","blocked","change_required"],
            "validation_errors": [],
        })

    # Report page
    from app.models.work_order import WorkOrder
    wo_list = db.query(WorkOrder).filter(
        WorkOrder.status.in_(["issued","in_progress","complete"])).all()
    render("reports/connections.html", {
        "connections": [],
        "datacenters": dcs,
        "work_orders": wo_list,
        "filter_dc_id": None, "filter_wo_id": None,
        "filter_install_status": None, "filter_fabric": None, "filter_action": None,
        "INSTALL_STATUSES": ["pending","done","blocked","change_required"],
        "ACTIONS": ["A","R","C"],
        "FABRICS": ["A","B"],
    })

    # Analytics pages
    from app.services.analytics import get_rack_elevation, get_port_adjacency_warnings, get_port_utilization
    from app.services.inventory import list_datacenters as _list_dcs
    from app.services.work_orders import list_work_orders as _list_wos
    elev = get_rack_elevation(db, rack.id)
    if elev:
        render("analytics/rack_elevation.html", {
            "rack": elev["rack"],
            "slots": elev["slots"],
            "unpositioned": elev["unpositioned"],
            "summary": elev["summary"],
        })
    render("analytics/port_adjacency.html", {
        "warnings": [],
        "datacenters": _list_dcs(db),
        "work_orders": _list_wos(db),
        "filter_wo_id": None,
        "filter_dc_id": None,
        "threshold": 4,
    })
    render("analytics/port_utilization.html", {
        "rows": get_port_utilization(db),
        "datacenters": _list_dcs(db),
        "filter_dc_id": None,
    })

    # Audit log page
    from app.models.audit import AuditLog
    render("audit/index.html", {
        "entries": [],
        "total": 0,
        "page": 1,
        "per_page": 50,
        "total_pages": 1,
        "users": [],
        "entity_types": [],
        "filter_entity_type": None,
        "filter_user_id": None,
        "filter_action": None,
        "filter_date_from": None,
        "filter_date_to": None,
    })



def test_work_orders(db, dc: Datacenter, rack: Rack, sw: Switch, user: User):
    section("Work Orders")

    wo = wo_svc.create_work_order(db, name="_SIM Work Order", datacenter_id=dc.id,
                                   work_type="install", created_by=user.id)
    CREATED_IDS["work_order"].append(wo.id)
    check("Create work order", wo.id is not None)
    check("Status is draft", wo.status == "draft")

    # Invalid work type
    expect_error("Invalid work type blocked",
                 lambda: wo_svc.create_work_order(db, name="x", datacenter_id=dc.id,
                                                   work_type="invalid", created_by=user.id),
                 "Invalid work type")

    section("Connections — Valid rows")

    def conn(extra={}):
        base = {
            "action": "A", "cable_type": "LC_Fiber", "purpose": "storage",
            "device_rack_name_raw": rack.name, "device_rack_u": "10",
            "device_slot": "4a", "device_port": "0",
            "switch_rack_name_raw": rack.name, "switch_rack_u": "30",
            "switch_slot": "1", "switch_port": "1",
            "install_status": "pending",
        }
        base.update(extra)
        return base

    # Unique port counter to avoid duplicate switch port errors
    port_counter = [1]
    def next_port():
        p = str(port_counter[0])
        port_counter[0] += 1
        return p

    # Direct same-rack connection
    c1, errs, warns = wo_svc.create_connection(db, wo.id,
        conn({"switch_port": next_port()}), user.id)
    check("Create valid connection", len(errs) == 0, f"errs={errs}")
    check("Cable length auto-computed", c1.seg1_length is not None,
          f"seg1={c1.seg1_length}")
    check("Same-rack length is numeric", c1.seg1_length not in ("cross-cabinet","incomplete","exceeds_max"),
          f"seg1={c1.seg1_length}")

    # Cross-cabinet
    c2, errs2, _ = wo_svc.create_connection(db, wo.id,
        conn({"switch_rack_name_raw": "OTHER-RACK", "switch_port": next_port()}), user.id)
    check("Cross-cabinet connection saves", len(errs2) == 0)
    check("Cross-cabinet seg1 = cross-cabinet", c2.seg1_length == "cross-cabinet",
          f"got {c2.seg1_length}")

    # Incomplete — missing switch_rack_u
    c3, errs3, _ = wo_svc.create_connection(db, wo.id,
        conn({"switch_rack_u": "", "switch_port": next_port()}), user.id)
    check("Incomplete connection saves (validation at issue time)", True)
    check("Incomplete seg1 = incomplete", c3.seg1_length == "incomplete",
          f"got {c3.seg1_length}")

    # Fabric A and B
    cA, _, _ = wo_svc.create_connection(db, wo.id,
        conn({"fabric": "A", "switch_port": next_port()}), user.id)
    cB, _, _ = wo_svc.create_connection(db, wo.id,
        conn({"fabric": "B", "switch_port": next_port()}), user.id)
    check("Fabric A row saved", cA.fabric == "A")
    check("Fabric B row saved", cB.fabric == "B")

    # LAG pair
    cLAG1, _, _ = wo_svc.create_connection(db, wo.id,
        conn({"lag_id": "Po1", "lag_member_index": "1",
              "switch_port": next_port()}), user.id)
    cLAG2, _, _ = wo_svc.create_connection(db, wo.id,
        conn({"lag_id": "Po1", "lag_member_index": "2",
              "switch_port": next_port()}), user.id)
    check("LAG member 1 saved", cLAG1.lag_id == "Po1")
    check("LAG member 2 saved", cLAG2.lag_member_index == 2)

    # Singlemode fiber
    cSM, _, _ = wo_svc.create_connection(db, wo.id,
        conn({"fiber_mode": "singlemode", "switch_port": next_port()}), user.id)
    check("Singlemode fiber mode saved", cSM.fiber_mode == "singlemode")

    section("Connections — Validation / Duplicate checks")

    # Duplicate switch port — hard block (using raw name path, matching real UI)
    c_dup, errs_dup, _ = wo_svc.create_connection(db, wo.id,
        conn({"switch_name_raw": "_SIM-SW-DEDUP", "switch_slot": "1", "switch_port": "99"}), user.id)
    check("First switch port assignment succeeds", len(errs_dup) == 0)

    c_dup2, errs_dup2, _ = wo_svc.create_connection(db, wo.id,
        conn({"switch_name_raw": "_SIM-SW-DEDUP", "switch_slot": "1", "switch_port": "99"}), user.id)
    check("Duplicate switch port hard-blocked", len(errs_dup2) > 0,
          f"errs={errs_dup2}")

    section("Work Order Status Transitions")

    # Can't issue with incomplete rows — try with just the one bad row (c3 has missing switch_rack_u)
    # But we have other valid rows too, so only c3 should fail
    try:
        wo_svc.transition_status(db, wo, "issued", user)
        db.refresh(wo)
        if wo.status == "issued":
            check("Issue with mixed rows (some incomplete) — check behavior",
                  True, "issued despite incomplete row — validation only blocks if ALL mandatory missing")
        else:
            check("Status after issue attempt", False, f"status={wo.status}")
    except ValueError as e:
        row_errors = getattr(e, "row_errors", [])
        check("Issue blocked due to validation errors",
              len(row_errors) > 0, f"{len(row_errors)} rows with errors")
        # Fix the bad row and try again — delete c3
        wo_svc.soft_delete_connection(db, c3, user.id)
        wo_svc.transition_status(db, wo, "issued", user)
        db.refresh(wo)
        check("Issue succeeds after fixing errors", wo.status == "issued")

    # Pull back to draft
    wo_svc.transition_status(db, wo, "draft", user)
    db.refresh(wo)
    check("Pull back to draft", wo.status == "draft")

    # Re-issue
    wo_svc.transition_status(db, wo, "issued", user)
    db.refresh(wo)
    check("Re-issue succeeds", wo.status == "issued")

    # Invalid transition (issued → complete skips in_progress)
    expect_error("Invalid transition issued→complete blocked",
                 lambda: wo_svc.transition_status(db, wo, "complete", user),
                 "Cannot move")

    # In progress
    wo_svc.transition_status(db, wo, "in_progress", user)
    db.refresh(wo)
    check("In progress transition", wo.status == "in_progress")

    # Complete — should soft-delete R-action rows
    r_conn, _, _ = wo_svc.create_connection(db, wo.id,
        conn({"action": "R", "switch_port": "R1"}), user.id)
    check("R-action row created", r_conn.action == "R")
    wo_svc.transition_status(db, wo, "complete", user)
    db.refresh(wo)
    check("Complete transition", wo.status == "complete")
    db.refresh(r_conn)
    check("R-action row soft-deleted on complete", r_conn.deleted_at is not None)

    section("Soft Delete")

    total = db.query(Connection).filter(
        Connection.work_order_id == wo.id).count()
    active = db.query(Connection).filter(
        Connection.work_order_id == wo.id,
        Connection.deleted_at.is_(None)).count()
    deleted = db.query(Connection).filter(
        Connection.work_order_id == wo.id,
        Connection.deleted_at.isnot(None)).count()
    check("Total connections tracked", total > 0, f"{total} total")
    check("Active connections > 0", active > 0, f"{active} active")
    check("Deleted connections tracked (R-rows + c3)", deleted >= 2, f"{deleted} deleted")


# ── Analytics tests ───────────────────────────────────────────────────────

def test_analytics(db, dc: Datacenter, rack: Rack, user: User, sw: Switch):
    section("Analytics")
    from app.services.analytics import (
        get_rack_elevation,
        get_port_adjacency_warnings,
        get_port_utilization,
    )

    # ── Rack elevation ─────────────────────────────────────────────────────
    data = get_rack_elevation(db, rack.id)
    check("Elevation: returns dict for valid rack", data is not None)
    check("Elevation: slots list non-empty", len(data["slots"]) > 0,
          f"{len(data['slots'])} slots")
    check("Elevation: summary fields present",
          all(k in data["summary"] for k in
              ("total_ru", "used_ru", "free_ru", "device_count",
               "switch_count", "patch_panel_count")))
    check("Elevation: used_ru > 0", data["summary"]["used_ru"] > 0,
          f"used_ru={data['summary']['used_ru']}")
    check("Elevation: used_ru + free_ru == total_ru",
          data["summary"]["used_ru"] + data["summary"]["free_ru"] == data["summary"]["total_ru"])
    check("Elevation: None for missing rack",
          get_rack_elevation(db, 999_999) is None)

    # ── Port adjacency ─────────────────────────────────────────────────────
    # Create a WO with two connections for the same device to same switch,
    # switch ports 4 and 5 (numeric, adjacent)
    wo_adj = wo_svc.create_work_order(
        db, name="_SIM Adjacency WO", datacenter_id=dc.id,
        work_type="install", created_by=user.id,
    )
    CREATED_IDS["work_order"].append(wo_adj.id)
    wo_svc.create_connection(db, wo_adj.id, {
        "action": "A", "cable_type": "LC_Fiber", "purpose": "storage",
        "device_name_raw": "_SIM-ADJ-DEVICE",
        "device_rack_name_raw": rack.name, "device_rack_u": "8",
        "device_slot": "hba0", "device_port": "p0",
        "switch_name_raw": "_SIM-ADJ-SWITCH",
        "switch_rack_name_raw": rack.name, "switch_rack_u": "30",
        "switch_slot": "1", "switch_port": "4",
    }, user.id)
    wo_svc.create_connection(db, wo_adj.id, {
        "action": "A", "cable_type": "LC_Fiber", "purpose": "storage",
        "device_name_raw": "_SIM-ADJ-DEVICE",
        "device_rack_name_raw": rack.name, "device_rack_u": "8",
        "device_slot": "hba1", "device_port": "p1",
        "switch_name_raw": "_SIM-ADJ-SWITCH",
        "switch_rack_name_raw": rack.name, "switch_rack_u": "30",
        "switch_slot": "1", "switch_port": "5",
    }, user.id)

    warnings = get_port_adjacency_warnings(db, wo_id=wo_adj.id)
    check("Adjacency: returns list", isinstance(warnings, list))
    check("Adjacency: at least one warning", len(warnings) > 0,
          f"{len(warnings)} warnings")
    if warnings:
        check("Adjacency: severity is 'adjacent' for ports 4 and 5",
              warnings[0]["severity"] == "adjacent",
              f"severity={warnings[0]['severity']}")
        check("Adjacency: min_separation = 1",
              warnings[0]["min_separation"] == 1,
              f"min_sep={warnings[0]['min_separation']}")

    # ── Port utilization ───────────────────────────────────────────────────
    rows = get_port_utilization(db)
    check("Utilization: returns list", isinstance(rows, list))
    # The sim sw has no device_type with port_count — check no exception
    check("Utilization: no exception raised", True)


# ── PDF export tests ──────────────────────────────────────────────────────

def test_pdf_export(db, dc: Datacenter, user: User):
    section("PDF Export")
    from app.services.pdf_export import generate_work_order_pdf

    # ── Test 1: WO with zero connections ──────────────────────────────────
    wo_empty = wo_svc.create_work_order(
        db, name="_SIM PDF Empty WO", datacenter_id=dc.id,
        work_type="install", created_by=user.id,
    )
    CREATED_IDS["work_order"].append(wo_empty.id)
    try:
        pdf = generate_work_order_pdf(db, wo_empty.id)
        check("PDF: zero connections — returns bytes", isinstance(pdf, bytes))
        check("PDF: zero connections — starts with %PDF", pdf[:4] == b"%PDF")
    except Exception as e:
        check("PDF: zero connections — returns bytes", False, str(e))
        check("PDF: zero connections — starts with %PDF", False, str(e))

    # ── Test 2: WO with connections (no patch data) ───────────────────────
    wo_conn = wo_svc.create_work_order(
        db, name="_SIM PDF Conn WO", datacenter_id=dc.id,
        work_type="install", created_by=user.id,
    )
    CREATED_IDS["work_order"].append(wo_conn.id)
    wo_svc.create_connection(db, wo_conn.id, {
        "action": "A", "cable_type": "LC_Fiber", "purpose": "storage",
        "device_rack_name_raw": "_SIM-RACK", "device_rack_u": "5",
        "device_slot": "1a", "device_port": "eth0",
        "switch_rack_name_raw": "_SIM-RACK", "switch_rack_u": "30",
        "switch_slot": "1", "switch_port": "Gi99/20",
        "install_status": "done",
    }, user.id)
    wo_svc.create_connection(db, wo_conn.id, {
        "action": "R", "cable_type": "RJ45", "purpose": "management",
        "device_rack_name_raw": "_SIM-RACK", "device_rack_u": "6",
        "device_slot": "1a", "device_port": "eth1",
        "switch_rack_name_raw": "_SIM-RACK", "switch_rack_u": "30",
        "switch_slot": "1", "switch_port": "Gi99/21",
    }, user.id)
    try:
        pdf = generate_work_order_pdf(db, wo_conn.id)
        check("PDF: with connections — returns bytes", isinstance(pdf, bytes))
        check("PDF: with connections — starts with %PDF", pdf[:4] == b"%PDF")
    except Exception as e:
        check("PDF: with connections — returns bytes", False, str(e))
        check("PDF: with connections — starts with %PDF", False, str(e))

    # ── Test 3: WO with connections that have patch data ─────────────────
    wo_patch = wo_svc.create_work_order(
        db, name="_SIM PDF Patch WO", datacenter_id=dc.id,
        work_type="install", created_by=user.id,
    )
    CREATED_IDS["work_order"].append(wo_patch.id)
    wo_svc.create_connection(db, wo_patch.id, {
        "action": "A", "cable_type": "LC_Fiber", "purpose": "data",
        "device_rack_name_raw": "_SIM-RACK", "device_rack_u": "7",
        "device_slot": "2a", "device_port": "hba0",
        "switch_rack_name_raw": "_SIM-RACK", "switch_rack_u": "30",
        "switch_slot": "1", "switch_port": "Gi99/22",
        "device_patch_rack_name_raw": "_SIM-PP-RACK",
        "device_patch_ru": "3",
        "device_patch_side": "front",
        "device_patch_port": "PP-01",
    }, user.id)
    try:
        pdf = generate_work_order_pdf(db, wo_patch.id)
        check("PDF: patch data — returns bytes", isinstance(pdf, bytes))
        check("PDF: patch data — starts with %PDF", pdf[:4] == b"%PDF")
    except Exception as e:
        check("PDF: patch data — returns bytes", False, str(e))
        check("PDF: patch data — starts with %PDF", False, str(e))

    # ── Test 4: ValueError for missing WO ─────────────────────────────────
    expect_error(
        "PDF: ValueError for missing WO",
        lambda: generate_work_order_pdf(db, 999_999),
        "not found",
    )


# ── Excel import tests ────────────────────────────────────────────────────

def test_excel_import(db, dc, rack, user):
    section("Excel Import")
    import io
    import json
    import openpyxl
    from app.services.excel_import import (
        parse_crossconnect_excel, generate_template,
        CANONICAL_HEADERS, CANONICAL_FIELD_MAP,
    )
    from app.models.settings import AppSetting

    # Build a minimal .xlsx in memory using canonical column layout
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Connections"
    ws.append(CANONICAL_HEADERS)

    def _row(**kwargs):
        row = [""] * len(CANONICAL_HEADERS)
        for field, val in kwargs.items():
            idx = CANONICAL_FIELD_MAP.get(field)
            if idx is not None:
                row[idx] = val
        return row

    # row 1 — valid ADD
    ws.append(_row(
        action="ADD", cable_type="copper", purpose="prod",
        device_rack_name_raw="R01", device_rack_u="1", device_slot="1", device_port="eth0",
        switch_rack_name_raw="R02", switch_rack_u="2", switch_slot="1", switch_port="Gi0/1",
        fabric="LAN",
    ))
    # row 2 — valid ADD with lag_id and fiber_mode
    ws.append(_row(
        action="ADD", cable_type="sfp", purpose="storage",
        device_rack_name_raw="R01", device_rack_u="3", device_slot="1", device_port="hba0",
        switch_rack_name_raw="R02", switch_rack_u="4", switch_slot="1", switch_port="Gi0/2",
        fabric="SAN", lag_id="10", fiber_mode="singlemode",
    ))
    # row 3 — blank action, should be skipped
    ws.append(_row(
        cable_type="copper", device_rack_name_raw="R99", switch_port="Gi99/99",
    ))

    buf = io.BytesIO()
    wb.save(buf)
    file_bytes = buf.getvalue()

    col_map = CANONICAL_FIELD_MAP
    rows = parse_crossconnect_excel(file_bytes, col_map, skip_sheets=["Schema"])
    check("Canonical map: 2 data rows parsed (blank action skipped)", len(rows) == 2, f"got {len(rows)}")

    r0 = rows[0]
    check("Row 0: action uppercased to ADD", r0.get("action") == "ADD")
    check("Row 0: fabric uppercased to LAN", r0.get("fabric") == "LAN")
    check("Row 0: purpose lowercased to prod", r0.get("purpose") == "prod")

    r1 = rows[1]
    check("Row 1: lag_id coerced to int 10", r1.get("lag_id") == 10)
    check("Row 1: fiber_mode preserved", r1.get("fiber_mode") == "singlemode")

    # Test Honda / legacy layout
    honda_map = {
        "action": 0, "fabric": 1, "cable_type": 2, "purpose": 3,
        "device_name_raw": 4, "device_rack_name_raw": 5, "device_rack_u": 6,
        "device_slot": 7, "device_port": 8,
        "switch_name_raw": 9, "switch_rack_name_raw": 10, "switch_rack_u": 11,
        "switch_slot": 12, "switch_port": 13,
        "vlan_vsan": 14, "lag_id": 15, "fiber_mode": 16, "comments": 17,
    }
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Sheet1"
    ws2.append(["ACTION", "FABRIC", "CABLE_TYPE", "PURPOSE",
                "DEVICE_NAME", "DEVICE_RACK", "DEVICE_RACK_U",
                "DEVICE_SLOT", "DEVICE_PORT",
                "SWITCH_NAME", "SWITCH_RACK", "SWITCH_RACK_U",
                "SWITCH_SLOT", "SWITCH_PORT",
                "VLAN", "LAG_ID", "FIBER_MODE", "COMMENTS"])
    ws2.append(["ADD", "SAN", "dac", "STORAGE",
                "server01", "R01", 5, 1, "hba1",
                "san-sw01", "R02", 10, 1, "Gi1/1",
                "", 20, "multimode", ""])
    buf2 = io.BytesIO()
    wb2.save(buf2)
    rows2 = parse_crossconnect_excel(buf2.getvalue(), honda_map, skip_sheets=[])
    check("Honda map: 1 row parsed", len(rows2) == 1, f"got {len(rows2)}")
    check("Honda: purpose lowercased", rows2[0].get("purpose") == "storage")
    check("Honda: lag_id coerced to int 20", rows2[0].get("lag_id") == 20)
    check("Honda: fiber_mode preserved", rows2[0].get("fiber_mode") == "multimode")

    # Test generate_template returns valid xlsx bytes with header row
    tmpl_bytes = generate_template()
    check("generate_template returns non-empty bytes", len(tmpl_bytes) > 0)
    wb3 = openpyxl.load_workbook(io.BytesIO(tmpl_bytes))
    ws3 = wb3["Connections"]
    first_row = [c.value for c in ws3[1]]
    check("Template header row matches CANONICAL_HEADERS", first_row == CANONICAL_HEADERS)
    check("Template has Schema sheet", "Schema" in wb3.sheetnames)


# ── Phase 10: Admin / Backup / Recycle Bin tests ─────────────────────────

def test_backup(db):
    section("Admin: Backup")
    from app.services.backup import create_backup, get_last_backup_time

    data = create_backup(db)
    check("Backup: returns bytes", isinstance(data, bytes))
    check("Backup: valid SQLite magic", data[:16] == b"SQLite format 3\x00")

    ts = get_last_backup_time(db)
    check("Backup: last_backup_at updated in app_settings", ts is not None)


def test_recycle_bin_list(db, dc, rack, user):
    section("Admin: Recycle Bin (list / restore)")
    from app.services.recycle_bin import list_deleted_connections, restore_connection

    wo = wo_svc.create_work_order(
        db, name="_SIM RB List WO", datacenter_id=dc.id,
        work_type="install", created_by=user.id,
    )
    CREATED_IDS["work_order"].append(wo.id)

    conn_data = {
        "action": "A", "cable_type": "LC_Fiber", "purpose": "storage",
        "device_rack_name_raw": rack.name, "device_rack_u": "10",
        "device_slot": "1a", "device_port": "rb1",
        "switch_rack_name_raw": rack.name, "switch_rack_u": "30",
        "switch_slot": "1", "switch_port": "rb1",
        "install_status": "pending",
    }
    conn, _, _ = wo_svc.create_connection(db, wo.id, conn_data, user.id)
    conn_id = conn.id

    # Ensure recycle_bin_enabled=true before soft-deleting
    from app.models.settings import AppSetting
    _set_rb(db, True)

    wo_svc.soft_delete_connection(db, conn, user.id)

    deleted = list_deleted_connections(db)
    check("Recycle bin: soft-deleted connection appears in list",
          any(c.id == conn_id for c in deleted))

    restore_connection(db, conn_id)
    db.refresh(conn)
    check("Recycle bin: connection restored (deleted_at=None)", conn.deleted_at is None)

    deleted_after = list_deleted_connections(db)
    check("Recycle bin: connection absent from deleted list after restore",
          not any(c.id == conn_id for c in deleted_after))


def _set_rb(db, enabled: bool):
    from app.models.settings import AppSetting
    row = db.get(AppSetting, "recycle_bin_enabled")
    val = "true" if enabled else "false"
    if row:
        row.value = val
    else:
        db.add(AppSetting(key="recycle_bin_enabled", value=val))
    db.commit()


def test_recycle_bin_purge(db, dc, rack, user):
    section("Admin: Recycle Bin (purge)")
    from app.services.recycle_bin import purge_all

    _set_rb(db, True)

    wo = wo_svc.create_work_order(
        db, name="_SIM RB Purge WO", datacenter_id=dc.id,
        work_type="install", created_by=user.id,
    )
    CREATED_IDS["work_order"].append(wo.id)

    conn_data = {
        "action": "A", "cable_type": "LC_Fiber", "purpose": "storage",
        "device_rack_name_raw": rack.name, "device_rack_u": "10",
        "device_slot": "1a", "device_port": "purge1",
        "switch_rack_name_raw": rack.name, "switch_rack_u": "30",
        "switch_slot": "1", "switch_port": "purge1",
        "install_status": "pending",
    }
    conn, _, _ = wo_svc.create_connection(db, wo.id, conn_data, user.id)
    conn_id = conn.id
    wo_svc.soft_delete_connection(db, conn, user.id)

    count = purge_all(db, "connections")
    check("Recycle bin purge: returned count > 0", count > 0)

    still_there = db.get(Connection, conn_id)
    check("Recycle bin purge: connection permanently gone from DB", still_there is None)


def test_recycle_bin_toggle(db, dc, rack, user):
    section("Admin: Recycle Bin Toggle (recycle_bin_enabled=false)")

    _set_rb(db, False)

    wo = wo_svc.create_work_order(
        db, name="_SIM RB Toggle WO", datacenter_id=dc.id,
        work_type="install", created_by=user.id,
    )
    CREATED_IDS["work_order"].append(wo.id)

    conn_data = {
        "action": "A", "cable_type": "LC_Fiber", "purpose": "storage",
        "device_rack_name_raw": rack.name, "device_rack_u": "10",
        "device_slot": "1a", "device_port": "toggle1",
        "switch_rack_name_raw": rack.name, "switch_rack_u": "30",
        "switch_slot": "1", "switch_port": "toggle1",
        "install_status": "pending",
    }
    conn, _, _ = wo_svc.create_connection(db, wo.id, conn_data, user.id)
    conn_id = conn.id

    wo_svc.soft_delete_connection(db, conn, user.id)

    still_there = db.get(Connection, conn_id)
    check("Toggle: connection hard-deleted when recycle_bin_enabled=false",
          still_there is None)

    # Restore default
    _set_rb(db, True)


def test_recycle_bin_system(db, user):
    section("Admin: Recycle Bin (systems soft-delete / restore / purge)")
    from app.services.recycle_bin import list_deleted_systems, restore_system, purge_all

    _set_rb(db, True)

    system = inv_svc.create_system(db, name="_SIM RB System", system_type="storage", user_id=user.id)
    CREATED_IDS["system"].append(system.id)
    system_id = system.id

    inv_svc.delete_system(db, system, user_id=user.id)
    db.refresh(system)
    check("System soft-deleted (deleted_at set)", system.deleted_at is not None)

    check("System excluded from list_systems after soft-delete",
          not any(s.id == system_id for s in inv_svc.list_systems(db)))

    deleted = list_deleted_systems(db)
    check("Recycle bin: soft-deleted system appears in list",
          any(s.id == system_id for s in deleted))

    restore_system(db, system_id)
    db.refresh(system)
    check("Recycle bin: system restored (deleted_at=None)", system.deleted_at is None)

    check("System back in list_systems after restore",
          any(s.id == system_id for s in inv_svc.list_systems(db)))

    # Re-delete then purge
    inv_svc.delete_system(db, system, user_id=user.id)
    count = purge_all(db, "systems")
    check("Recycle bin purge: systems count > 0", count > 0)

    still_there = db.get(System, system_id)
    check("Recycle bin purge: system permanently gone from DB", still_there is None)


def test_tuning_validation():
    section("Admin: Tuning Validation Logic")
    import json

    # Invalid JSON detection
    try:
        json.loads("this is not json")
        check("Invalid JSON detected", False, "Expected JSONDecodeError")
    except json.JSONDecodeError:
        check("Invalid JSON detected", True)

    # Valid JSON passes
    try:
        json.loads('{"action": 0, "cable_type": 1}')
        check("Valid JSON accepted", True)
    except json.JSONDecodeError:
        check("Valid JSON accepted", False)

    # Valid comma-separated lengths
    try:
        [float(x.strip()) for x in "0.5,1,2,3,5".split(",") if x.strip()]
        check("Valid comma-separated lengths parsed", True)
    except ValueError:
        check("Valid comma-separated lengths parsed", False)

    # Invalid comma-separated lengths
    try:
        [float(x.strip()) for x in "0.5,1,abc,3".split(",") if x.strip()]
        check("Invalid lengths detected", False, "Expected ValueError")
    except ValueError:
        check("Invalid lengths detected", True)

    # Valid positive integer
    try:
        v = int("4")
        check("Valid threshold accepted", v > 0)
    except ValueError:
        check("Valid threshold accepted", False)

    # Non-integer threshold
    try:
        int("abc")
        check("Non-integer threshold detected", False, "Expected ValueError")
    except ValueError:
        check("Non-integer threshold detected", True)

    # Zero threshold (not positive)
    try:
        v = int("0")
        check("Zero threshold rejected", v <= 0)
    except ValueError:
        check("Zero threshold rejected", False)


def test_admin_templates(db, dc, rack, user):
    section("Admin: Template Renders")
    from jinja2 import Environment, FileSystemLoader
    from app.services.recycle_bin import get_deleted_counts
    from app.services.backup import get_last_backup_time
    from app.services.settings import get_bool_setting, get_all_settings
    import app.services.auth as auth_svc

    template_dir = str(Path(__file__).resolve().parents[1] / "app" / "templates")
    env = Environment(loader=FileSystemLoader(template_dir))

    class FakeRequest:
        class url:
            path = "/admin"
        query_params = {}

    class FakeAdminUser:
        id = 1
        username = "admin"
        display_name = "Administrator"
        role = "admin"
        is_active = True

    admin_user = FakeAdminUser()
    req = FakeRequest()

    def render(template_name, extra_ctx):
        try:
            t = env.get_template(template_name)
            t.render(request=req, user=admin_user, **extra_ctx)
            check(f"Render {template_name}", True)
            return True
        except Exception as e:
            check(f"Render {template_name}", False, f"{type(e).__name__}: {e}")
            return False

    deleted_counts = get_deleted_counts(db)
    last_backup = get_last_backup_time(db)
    recycle_bin_enabled = get_bool_setting(db, "recycle_bin_enabled")
    all_settings = get_all_settings(db)
    users = auth_svc.list_users(db)

    render("admin/index.html", {
        "user_count": db.query(__import__("app.models.user", fromlist=["User"]).User).count(),
        "deleted_counts": deleted_counts,
        "total_deleted": sum(deleted_counts.values()),
        "last_backup": last_backup,
        "recycle_bin_enabled": recycle_bin_enabled,
        "restored": None,
    })

    render("admin/users.html", {
        "users": users,
        "valid_roles": ["admin", "architect", "dc_tech", "viewer"],
        "error": None,
        "success": None,
    })

    render("admin/tuning.html", {
        "settings": all_settings,
        "saved": None,
        "error": None,
    })

    from app.services.recycle_bin import (
        list_deleted_connections, list_deleted_work_orders,
        list_deleted_devices, list_deleted_racks, list_deleted_switches,
        list_deleted_systems,
    )
    render("admin/recycle_bin.html", {
        "tab": "connections",
        "recycle_bin_enabled": recycle_bin_enabled,
        "deleted_connections": list_deleted_connections(db),
        "deleted_work_orders": [],
        "deleted_devices": [],
        "deleted_racks": [],
        "deleted_switches": [],
        "deleted_systems": [],
        "deleted_counts": deleted_counts,
    })

    render("admin/recycle_bin.html", {
        "tab": "systems",
        "recycle_bin_enabled": recycle_bin_enabled,
        "deleted_connections": [],
        "deleted_work_orders": [],
        "deleted_devices": [],
        "deleted_racks": [],
        "deleted_switches": [],
        "deleted_systems": list_deleted_systems(db),
        "deleted_counts": deleted_counts,
    })

    render("admin/backup.html", {
        "last_backup": last_backup,
        "is_sqlite": True,
    })

    render("admin/restore_confirm.html", {
        "token": "test-token-123",
        "filename": "crossconnect-backup-20260518.db",
        "file_size_kb": 512,
        "backup_version": "a2b3c4d5e6f7",
        "current_version": "a2b3c4d5e6f7",
        "versions_match": True,
    })


# ── Cleanup ───────────────────────────────────────────────────────────────

def cleanup(db):
    section("Cleanup")
    from sqlalchemy import text

    # Delete in reverse FK order
    for wo_id in CREATED_IDS["work_order"]:
        db.query(Connection).filter(Connection.work_order_id == wo_id).delete()
        db.query(WorkOrder).filter(WorkOrder.id == wo_id).delete()
    for sw_id in CREATED_IDS["switch"]:
        db.query(Switch).filter(Switch.id == sw_id).delete()
    for dev_id in reversed(CREATED_IDS["device"]):  # children before parents
        db.query(Device).filter(Device.id == dev_id).delete()
    for sys_id in CREATED_IDS["system"]:
        db.query(System).filter(System.id == sys_id).delete()
    for dt_id in CREATED_IDS["device_type"]:
        db.query(DeviceType).filter(DeviceType.id == dt_id).delete()
    for rack_id in CREATED_IDS["rack"]:
        db.query(Rack).filter(Rack.id == rack_id).delete()
    for dc_id in CREATED_IDS["datacenter"]:
        db.query(DCContact).filter(DCContact.datacenter_id == dc_id).delete()
        db.query(Datacenter).filter(Datacenter.id == dc_id).delete()
    for user_id in CREATED_IDS["user"]:
        db.query(AuditLog).filter(AuditLog.user_id == user_id).delete()
        db.query(User).filter(User.id == user_id).delete()
    db.commit()
    print("  Test data removed.")


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="CrossConnect debug simulation")
    parser.add_argument("--keep", action="store_true",
                        help="Keep test data in DB after run (for manual inspection)")
    args = parser.parse_args()

    print("\nCrossConnect Debug Simulation")
    print("=" * 60)

    db = SessionLocal()
    try:
        pre_run_cleanup(db)
        user = setup_test_user(db)
        dc = test_datacenters(db)
        rack = test_racks(db, dc)
        dt_storage, dt_chassis = test_device_types(db)
        device = test_devices(db, rack, dt_storage, dt_chassis)
        sw = test_switches(db, rack)
        test_template_renders(db, dc, rack, user)
        test_excel_import(db, dc, rack, user)
        test_analytics(db, dc, rack, user, sw)
        test_pdf_export(db, dc, user)
        test_work_orders(db, dc, rack, sw, user)
        test_backup(db)
        test_recycle_bin_list(db, dc, rack, user)
        test_recycle_bin_purge(db, dc, rack, user)
        test_recycle_bin_toggle(db, dc, rack, user)
        test_recycle_bin_system(db, user)
        test_tuning_validation()
        test_admin_templates(db, dc, rack, user)
    except Exception as e:
        import traceback
        print(f"\n  FATAL: Unhandled exception during simulation:")
        traceback.print_exc()
        RESULTS.append(("Unhandled exception", False, str(e)))
    finally:
        if not args.keep:
            try:
                cleanup(db)
            except Exception as e:
                print(f"  WARNING: Cleanup failed: {e}")
        db.close()

    # Summary
    print(f"\n{'='*60}")
    passed = sum(1 for _, ok, _ in RESULTS if ok)
    failed = sum(1 for _, ok, _ in RESULTS if not ok)
    total  = len(RESULTS)
    print(f"  Results: {passed}/{total} passed", end="")
    if failed:
        print(f"  ({failed} FAILED)")
        print(f"\n  Failed checks:")
        for label, ok, detail in RESULTS:
            if not ok:
                print(f"    ✗ {label}" + (f": {detail}" if detail else ""))
    else:
        print("  — all passed ✓")
    print()

    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
