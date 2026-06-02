"""
Label export service — generates a Brother P-touch Editor mail-merge .xlsx.

build_label_rows(db, wo_id) -> list[dict]
export_labels_xlsx(db, wo_id) -> bytes
"""
import re
from io import BytesIO

import openpyxl
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.services.work_orders import list_connections


def _s(val) -> str:
    """None-safe string — empty string instead of 'None'."""
    return "" if val is None else str(val)


def _san_line1(c) -> str:
    return (
        f"{_s(c.device_name_raw)} : {_s(c.device_grid)} : "
        f"U{_s(c.device_rack_u)} : C{_s(c.device_slot)} : P{_s(c.device_port)}"
        f" -- M{_s(c.device_patch_module)} : P{_s(c.device_patch_port)}"
    )


def _san_line2(c) -> str:
    return (
        f"{_s(c.switch_patch_rack_name_raw)} : {_s(c.switch_name_raw)} : "
        f"BL{_s(c.switch_slot)} : P{_s(c.switch_port)}"
        f" -- SH{_s(c.switch_patch_module)} : P{_s(c.switch_patch_port)}"
    )


def _lan_line1(c) -> str:
    dev_side = (
        f"{_s(c.device_name_raw)} : {_s(c.device_grid)} : "
        f"U {_s(c.device_rack_u)}: C {_s(c.device_slot)} : P {_s(c.device_port)}"
    )
    # Build switch group from non-empty parts; omit missing fields and their separators.
    sw_parts = [p for p in [
        _s(c.switch_patch_rack_name_raw),
        _s(c.switch_name_raw),
        f"P {_s(c.switch_port)}" if _s(c.switch_port) else "",
    ] if p]
    return f"{dev_side} - ({', '.join(sw_parts)})"


_HEADERS = [
    "Label Type", "Fabric", "Device Name", "Device Grid", "Device U",
    "Device Slot", "Device Port", "Dev Patch Module", "Dev Patch Port",
    "Switch Patch Rack", "Switch Name", "Switch Slot", "Switch Port",
    "SW Patch Module", "SW Patch Port", "Line 1", "Line 2",
]

_ROW_KEYS = [
    "label_type", "fabric", "device_name", "device_grid", "device_rack_u",
    "device_slot", "device_port", "device_patch_module", "device_patch_port",
    "switch_patch_rack", "switch_name", "switch_slot", "switch_port",
    "switch_patch_module", "switch_patch_port", "line1", "line2",
]


def build_label_rows(db: Session, wo_id: int) -> list[dict]:
    connections = [
        c for c in list_connections(db, wo_id)
        if c.action != "R"
    ]
    rows = []
    for c in connections:
        is_san = c.purpose == "storage"
        label_type = "SAN" if is_san else "LAN"
        line1 = _san_line1(c) if is_san else _lan_line1(c)
        line2 = _san_line2(c) if is_san else ""
        rows.append({
            "label_type": label_type,
            "fabric": _s(c.fabric),
            "device_name": _s(c.device_name_raw),
            "device_grid": _s(c.device_grid),
            "device_rack_u": _s(c.device_rack_u),
            "device_slot": _s(c.device_slot),
            "device_port": _s(c.device_port),
            "device_patch_module": _s(c.device_patch_module),
            "device_patch_port": _s(c.device_patch_port),
            "switch_patch_rack": _s(c.switch_patch_rack_name_raw),
            "switch_name": _s(c.switch_name_raw),
            "switch_slot": _s(c.switch_slot),
            "switch_port": _s(c.switch_port),
            "switch_patch_module": _s(c.switch_patch_module),
            "switch_patch_port": _s(c.switch_patch_port),
            "line1": line1,
            "line2": line2,
        })
    return rows


def export_labels_xlsx(db: Session, wo_id: int) -> bytes:
    rows = build_label_rows(db, wo_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Labels"

    # Header row
    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.number_format = "@"
    ws.freeze_panes = "A2"

    # Data rows
    for row in rows:
        ws.append([row[k] for k in _ROW_KEYS])

    # Force TEXT format and collect max widths
    col_widths = [len(h) for h in _HEADERS]
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, key in enumerate(_ROW_KEYS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.number_format = "@"
            val_len = len(str(cell.value) if cell.value else "")
            if val_len > col_widths[c_idx - 1]:
                col_widths[c_idx - 1] = val_len

    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(width + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
