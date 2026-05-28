import io
import json
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

CANONICAL_HEADERS = [
    "ACTION", "FABRIC", "PORT_DESCRIPTION", "CABLE_TYPE", "PURPOSE",
    "SYSTEM_NAME_RAW", "DEVICE_NAME_RAW", "DEVICE_SERIAL",
    "DEVICE_RACK_NAME_RAW", "DEVICE_GRID", "DEVICE_RACK_U", "DEVICE_SLOT", "DEVICE_PORT",
    "DEVICE_PATCH_RACK_NAME_RAW", "DEVICE_PATCH_RU", "DEVICE_PATCH_SIDE",
    "DEVICE_PATCH_MODULE", "DEVICE_PATCH_PORT",
    "SWITCH_PATCH_RACK_NAME_RAW", "SWITCH_PATCH_RU", "SWITCH_PATCH_SIDE",
    "SWITCH_PATCH_MODULE", "SWITCH_PATCH_PORT",
    "SWITCH_NAME_RAW", "SWITCH_SERIAL", "SWITCH_RACK_NAME_RAW",
    "SWITCH_GRID", "SWITCH_RACK_U", "SWITCH_SLOT", "SWITCH_PORT",
    "VLAN_VSAN", "LAG_ID", "FIBER_MODE", "COMMENTS",
]

CANONICAL_FIELD_MAP = {h.lower(): i for i, h in enumerate(CANONICAL_HEADERS)}

_INT_FIELDS = {"device_rack_u", "device_patch_ru", "switch_rack_u", "switch_patch_ru", "lag_id"}
_UPPER_FIELDS = {"action", "fabric"}
_LOWER_FIELDS = {"purpose"}
_NONE_IF_BLANK = {"cable_type", "fiber_mode"}


def _coerce(field: str, raw: Any) -> str | int | None:
    if raw is None:
        return None
    val = str(raw).strip()
    if not val or val == "-":
        return None if field in _NONE_IF_BLANK else ""
    if field in _UPPER_FIELDS:
        return val.upper()
    if field in _LOWER_FIELDS:
        return val.lower()
    if field in _INT_FIELDS:
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return None
    return val


def parse_crossconnect_excel(
    file_bytes: bytes,
    column_map: dict,
    skip_sheets: list[str],
) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    rows_out = []
    for sheet in wb.worksheets:
        if sheet.title in skip_sheets:
            continue
        effective_map: dict | None = None
        for row in sheet.iter_rows(values_only=True):
            if all(v is None for v in row):
                continue
            if effective_map is None:
                # Auto-detect column positions from header row; explicit
                # column_map overrides any auto-detected positions.
                auto = {
                    str(v).strip().lower(): i
                    for i, v in enumerate(row)
                    if v is not None
                }
                effective_map = {**auto, **column_map}
                continue
            if all(v is None for v in row):
                continue
            record: dict = {}
            for field, col_idx in effective_map.items():
                if col_idx < len(row):
                    record[field] = _coerce(field, row[col_idx])
                else:
                    record[field] = None
            action = record.get("action")
            if not action:
                continue
            rows_out.append(record)
    return rows_out


def generate_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Connections"

    header_fill = PatternFill("solid", fgColor="1A2A3A")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    required_fill = PatternFill("solid", fgColor="0D3349")

    required = {
        "ACTION", "CABLE_TYPE", "PURPOSE",
        "DEVICE_RACK_NAME_RAW", "DEVICE_RACK_U", "DEVICE_SLOT", "DEVICE_PORT",
        "SWITCH_RACK_NAME_RAW", "SWITCH_RACK_U", "SWITCH_SLOT", "SWITCH_PORT",
    }

    for col_idx, header in enumerate(CANONICAL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = required_fill if header in required else header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(12, len(header) + 2)

    # Force VLAN_VSAN column to text format so Excel preserves comma-separated
    # values like "100,101" instead of coercing them to integer 100101.
    _vlan_col_letter = ws.cell(row=1, column=CANONICAL_HEADERS.index('VLAN_VSAN') + 1).column_letter
    ws.column_dimensions[_vlan_col_letter].number_format = '@'

    dv_action = DataValidation(
        type="list",
        formula1='"ADD,REMOVE,MOVE,CONFIRM"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv_action)
    dv_action.sqref = f"A2:A1048576"

    schema_ws = wb.create_sheet("Schema")
    schema_ws["A1"] = "Field"
    schema_ws["B1"] = "Required"
    schema_ws["C1"] = "Notes"
    schema_ws["A1"].font = Font(bold=True)
    schema_ws["B1"].font = Font(bold=True)
    schema_ws["C1"].font = Font(bold=True)
    notes = {
        "ACTION": "ADD / REMOVE / MOVE / CONFIRM (uppercase)",
        "FABRIC": "LAN / SAN / OOB / etc. (uppercase)",
        "CABLE_TYPE": "copper / sfp / dac / etc.",
        "PURPOSE": "prod / mgmt / storage / etc. (lowercase)",
        "DEVICE_RACK_U": "Integer",
        "SWITCH_RACK_U": "Integer",
        "LAG_ID": "Integer port-channel / LAG number",
        "FIBER_MODE": "singlemode or multimode",
    }
    for i, header in enumerate(CANONICAL_HEADERS, start=2):
        schema_ws.cell(row=i, column=1, value=header)
        schema_ws.cell(row=i, column=2, value="YES" if header in required else "")
        schema_ws.cell(row=i, column=3, value=notes.get(header, ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
